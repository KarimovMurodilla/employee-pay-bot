from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def write_revenue_excel(data: dict, filename="revenue_summary.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue Summary"

    # Заголовок
    ws.merge_cells("A1:B1")
    ws["A1"] = f"Revenue Summary for {data['name']}"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Данные
    rows = [
        ("Establishment ID", data["establishment_id"]),
        ("Name", data["name"]),
        ("Total Revenue", f"${data['total_revenue']:.2f}"),
        ("Today Revenue", f"${data['today_revenue']:.2f}"),
        ("Total Orders", data["total_orders"]),
        ("Average Order Value", f"${data['average_order_value']:.2f}"),
    ]

    start_row = 3
    for i, (key, value) in enumerate(rows, start=start_row):
        ws[f"A{i}"] = key
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"A{i}"].fill = PatternFill(
            start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
        )
        ws[f"B{i}"] = value

    # Автоматическая ширина колонок
    for col in ["A", "B"]:
        ws.column_dimensions[col].width = 25

    wb.save(filename)
    return True


# data = {
#     "establishment_id": 1,
#     "name": "Cafe Mocha",
#     "total_revenue": 1200.50,
#     "today_revenue": 150.75,
#     "total_orders": 30,
#     "average_order_value": 40.02,
# }
# # Пример использования
# write_revenue_excel(data, filename=f"reports/revenue-summary_{datetime.now().strftime('%d-%m-%Y %H-%M-%S')}.xlsx") # revenue summary name with date
